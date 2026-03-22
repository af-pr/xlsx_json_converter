"""
XLSX parsing module for converting binary data to structured data.
"""

from typing import List
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import SheetData, Cell, DataType
from exceptions import InvalidFormatError


def parse(file_bytes: bytes) -> List[SheetData]:
    """
    Parse XLSX binary content into SheetData structures.
    
    Extracts all sheets from workbook using openpyxl. The first row in each sheet
    is ALWAYS treated as column headers. Empty header cells are auto-generated as
    Column_1, Column_2, etc. Each cell is represented with its original data type
    preserved if valid (string, number, date, boolean, formula, error);
    unrecognized types are cast to string, and empty cells are assigned None type.
    
    Input Format Requirement:
    - First row MUST contain column header names
    - Empty cells in first row will auto-generate Column_N names
    - Data rows start from row 2 onwards
    
    Args:
        file_bytes: Binary XLSX file content
        
    Returns:
        List of SheetData objects (one per sheet)
        
    Raises:
        InvalidFormatError: If file is corrupted or not valid XLSX
    """
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=False)
    except Exception as e:
        raise InvalidFormatError(f"Cannot parse XLSX file: {str(e)}")
    
    sheet_data_list = []
    for sheet in workbook.worksheets:
        headers = _extract_headers(sheet)
        rows = _extract_data(sheet, len(headers))
        sheet_data_list.append(SheetData(name=sheet.title, headers=headers, rows=rows))
    
    return sheet_data_list

def _extract_headers(worksheet : Worksheet) -> List[str]:
    """
    Extract first row of the sheet as column headers
    If any header cell is empty, auto-generate as Column_N
    where n is the 1-based column index (e.g., Column_1, Column_2, etc.)
    
    Args:
        sheet: openpyxl worksheet object
    Returns:
        List of header names
    """
    first_row = worksheet[1]  # This is the first row (headers), openpyxl starts indexing at 1
    headers = []
    for cell in first_row:
        if cell.value:
            headers.append(str(cell.value))
        else:
            headers.append(f"Column_{cell.column}")
    return headers

def _extract_data(worksheet: Worksheet, num_columns: int) -> List[List[Cell]]:
    """
    Extract data rows (from row 2 onwards)
    
    Args:
        worksheet: openpyxl worksheet object
        num_columns: Number of columns based on headers (to handle missing cells)   
    Returns:
        List of rows, where each row is a List[Cell]
    """
    rows = []
    
    # Iterate from row 2 onwards. If the sheet has fewer columns than 2, the loop won't be executed (empty content case)
    for row_idx in range(2, worksheet.max_row + 1):
        row = worksheet[row_idx]
        cell_list = []
        
        for cell in row:
            if cell.value is None:
                cell_type = None
                cell_val = None
            else:
                try:
                    cell_type = DataType(cell.data_type)
                    cell_val = cell.value
                except ValueError:
                    cell_type = DataType.STRING
                    cell_val = str(cell.value)
            
            cell_list.append(
                Cell(data_type=cell_type, value=cell_val)
            )
        
        # If current row has fewer cells than num_columns, add empty cells to maintain structure
        if len(cell_list) < num_columns:
            padding_needed = num_columns - len(cell_list)
            cell_list.extend([Cell(None, None) for _ in range(padding_needed)])
        
        rows.append(cell_list)
    
    return rows
